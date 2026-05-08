from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import scipy.io
import torch.utils.data as data
from openpyxl import load_workbook
from PIL import Image

from .splits import canonical_dataset_name
from .transforms import build_iqa_transform


def pil_loader(path) -> Image.Image:
    with open(path, "rb") as file:
        image = Image.open(file)
        return image.convert("RGB")


def list_files_by_suffix(path, suffix: str) -> List[str]:
    return [name for name in os.listdir(path) if os.path.splitext(name)[1] == suffix]


def list_tid_reference_ids(path, suffixes: str) -> List[str]:
    ids = []
    for name in os.listdir(path):
        if suffixes.find(os.path.splitext(name)[1]) != -1:
            ids.append(name[1:3])
    return ids


def scale_to_range(
    values,
    min_val: float = 0,
    max_val: float = 100,
    target_min: float = -50,
    target_max: float = 150,
):
    values = np.asarray(values)
    return (values - min_val) / (max_val - min_val) * (target_max - target_min) + target_min


class IQADataset(data.Dataset):
    def __init__(self, samples: Sequence[Tuple[str, float]], transform: Optional[Callable] = None) -> None:
        self.samples = list(samples)
        self.transform = transform

    def __getitem__(self, index: int) -> Dict[str, Any]:
        path, target = self.samples[index]
        image = pil_loader(path)
        if self.transform is not None:
            image = self.transform(image)
        return {
            "image": image,
            "target": np.asarray(target).astype(np.float32),
            "filename": os.path.basename(path),
        }

    def __len__(self) -> int:
        return len(self.samples)


class GeneratedDataset(IQADataset):
    """Generic generated IQA dataset from label CSV files."""

    def __init__(self, root, index, transform, label_path, patch_num: int = 1):
        label_dir = Path(label_path)
        label_files = sorted(label_dir.glob("*.csv"))
        samples = []
        for file_index in index:
            label_file = label_files[file_index]
            label_df = pd.read_csv(label_file)
            image_column = _first_existing_column(label_df, ["distorted_image", "dist_img", "image", "filename"])
            score_column = _first_existing_column(label_df, ["quality_score", "label", "mos", "dmos"])
            for row in label_df.itertuples(index=False):
                image_name = getattr(row, image_column)
                score = float(getattr(row, score_column))
                for _ in range(patch_num):
                    samples.append((str(Path(root) / image_name), score))
        super().__init__(samples, transform)


class SAQTIQADataset(IQADataset):
    """SAQT-IQA dataset built from generated labels and semantic similarity metadata."""

    def __init__(
        self,
        pretrain_path: str,
        index: Iterable[int],
        transform: Callable,
        label_path: str,
        similarity_csv_path: str,
        patch_num: int = 1,
        target_distortion_types: Optional[Iterable[str]] = None,
    ) -> None:
        self.pretrain_path = Path(pretrain_path)
        self.label_path = Path(label_path)
        self.similarity_csv_path = Path(similarity_csv_path)
        self.target_distortion_types = (
            {str(item).zfill(2) for item in target_distortion_types}
            if target_distortion_types is not None
            else None
        )
        path_mapping = self._build_image_path_mapping()
        selected_refs = {f"I{idx + 1:02d}.png" for idx in index}
        selected_pristine_images = self._load_selected_pristine_images(selected_refs)
        samples = self._load_label_samples(path_mapping, selected_pristine_images, patch_num)
        super().__init__(samples, transform)

    def _build_image_path_mapping(self) -> Dict[str, str]:
        mapping = {}
        for image_path in self.pretrain_path.rglob("*.bmp"):
            mapping[image_path.name] = str(image_path)
        if not mapping:
            raise ValueError(f"No .bmp images found under {self.pretrain_path}")
        return mapping

    def _load_selected_pristine_images(self, selected_refs: set[str]) -> set[str]:
        similarity_df = pd.read_csv(self.similarity_csv_path)
        required = {"kadis", "kadid-10k"}
        missing = required - set(similarity_df.columns)
        if missing:
            raise ValueError(f"{self.similarity_csv_path} is missing columns: {sorted(missing)}")
        filtered = similarity_df[similarity_df["kadid-10k"].isin(selected_refs)]
        return set(filtered["kadis"].dropna().astype(str).tolist())

    def _load_label_samples(
        self,
        path_mapping: Dict[str, str],
        selected_pristine_images: set[str],
        patch_num: int,
    ) -> List[Tuple[str, float]]:
        label_files = sorted(self.label_path.glob("*.csv"))
        if not label_files:
            raise FileNotFoundError(f"No label CSV files found in {self.label_path}")

        samples = []
        missing_images = []
        for label_file in label_files:
            label_df = pd.read_csv(label_file)
            image_column = _first_existing_column(label_df, ["distorted_image", "dist_img", "image", "filename"])
            score_column = _first_existing_column(label_df, ["quality_score", "label", "mos", "dmos"])
            for row in label_df.itertuples(index=False):
                image_name = str(getattr(row, image_column)).replace("\\", "/")
                filename = image_name.split("/")[-1]
                ref_name, dist_type = _parse_generated_distortion_name(filename)
                if ref_name not in selected_pristine_images:
                    continue
                if self.target_distortion_types is not None and dist_type not in self.target_distortion_types:
                    continue
                if filename not in path_mapping:
                    missing_images.append(filename)
                    continue
                score = float(getattr(row, score_column))
                for _ in range(patch_num):
                    samples.append((path_mapping[filename], score))

        if not samples:
            raise ValueError(
                "No SAQT-IQA samples loaded. Check label_path, similarity_csv_path, "
                "selected indices, and generated image paths."
            )
        return samples




class BID(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        workbook = load_workbook(root / "DatabaseGrades.xlsx")
        sheet = workbook.active
        image_names, mos_values = [], []
        for row_idx in range(2, 588):
            image_num = sheet.cell(row=row_idx, column=1).value
            mos = np.asarray(sheet.cell(row=row_idx, column=2).value).astype(np.float32)
            image_names.append(f"DatabaseImage{image_num:04d}.JPG")
            mos_values.append(mos)
        samples = _repeat_samples(root, image_names, mos_values, index, patch_num)
        super().__init__(samples, transform)


class SPAQ(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        info = pd.read_excel(root / "Annotations" / "MOS and Image attribute scores.xlsx", engine="openpyxl")
        image_names = np.asarray(info["Image name"])
        labels = np.asarray(info["MOS"]).astype(np.float32)
        samples = []
        for item in index:
            for _ in range(patch_num):
                samples.append((str(root / "TestImage" / image_names[item]), labels[item]))
        super().__init__(samples, transform)


class LIVEChallenge(IQADataset):
    def __init__(self, root, index, transform, patch_num, is_train):
        root = Path(root)
        image_mat = scipy.io.loadmat(root / "Data" / "AllImages_release.mat")
        image_names = image_mat["AllImages_release"][7:1169]
        mos_mat = scipy.io.loadmat(root / "Data" / "AllMOS_release.mat")
        labels = mos_mat["AllMOS_release"].astype(np.float32)[0][7:1169]
        if is_train:
            labels = scale_to_range(labels)
        samples = []
        for item in index:
            for _ in range(patch_num):
                samples.append((str(root / "Images" / image_names[item][0][0]), labels[item]))
        super().__init__(samples, transform)


class KonIQ10k(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        image_names, mos_values = [], []
        with open(root / "koniq10k_scores_and_distributions.csv", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                image_names.append(row["image_name"])
                mos_values.append(np.asarray(float(row["MOS_zscore"])).astype(np.float32))
        samples = []
        for item in index:
            for _ in range(patch_num):
                samples.append((str(root / "512x384" / image_names[item]), mos_values[item]))
        super().__init__(samples, transform)


class FBLIVE(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        image_names, mos_values = [], []
        with open(root / "labels_image.csv", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                image_names.append(row["name"])
                mos_values.append(np.asarray(float(row["mos"])).astype(np.float32))
        samples = _repeat_samples(root, image_names, mos_values, index, patch_num)
        super().__init__(samples, transform)


class LIVE(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        ref_names = sorted(list_files_by_suffix(root / "refimgs", ".bmp"))
        image_paths = (
            _live_distortion_paths(root / "jp2k", 227)
            + _live_distortion_paths(root / "jpeg", 233)
            + _live_distortion_paths(root / "wn", 174)
            + _live_distortion_paths(root / "gblur", 174)
            + _live_distortion_paths(root / "fastfading", 174)
        )
        dmos = scipy.io.loadmat(root / "dmos_realigned.mat")
        labels = dmos["dmos_new"].astype(np.float32)
        orgs = dmos["orgs"]
        ref_names_all = scipy.io.loadmat(root / "refnames_all.mat")["refnames_all"]

        samples = []
        for item in index:
            selected = (ref_names[item] == ref_names_all) * ~orgs.astype(np.bool_)
            selected_indices = np.where(selected == True)[1].tolist()
            for selected_idx in selected_indices:
                for _ in range(patch_num):
                    samples.append((image_paths[selected_idx], labels[0][selected_idx]))
        super().__init__(samples, transform)


class CSIQ(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        ref_names = sorted(list_files_by_suffix(root / "src_imgs", ".png"), reverse=True)
        image_names, targets, ref_names_all = [], [], []
        with open(root / "csiq_label.txt") as file:
            for line in file:
                words = line.strip().split()
                image_names.append(words[0])
                targets.append(words[1])
                ref_parts = words[0].split(".")
                ref_names_all.append(ref_parts[0] + "." + ref_parts[-1])
        labels = np.asarray(targets).astype(np.float32)
        ref_names_all = np.asarray(ref_names_all)

        samples = []
        for item in index:
            selected_indices = np.where(ref_names[item] == ref_names_all)[0].tolist()
            for selected_idx in selected_indices:
                for _ in range(patch_num):
                    samples.append((str(root / "all_dis_imgs" / image_names[selected_idx]), labels[selected_idx]))
        super().__init__(samples, transform)


class TID2013(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        ref_ids = sorted(list_tid_reference_ids(root / "reference_images", ".bmp.BMP"))
        image_names, targets, ref_ids_all = [], [], []
        with open(root / "mos_with_names.txt") as file:
            for line in file:
                words = line.strip().split()
                targets.append(words[0])
                image_names.append(words[1])
                ref_ids_all.append(words[1].split("_")[0][1:])
        labels = np.asarray(targets).astype(np.float32)
        ref_ids_all = np.asarray(ref_ids_all)
        samples = []
        for item in index:
            selected_indices = np.where(ref_ids[item] == ref_ids_all)[0].tolist()
            for selected_idx in selected_indices:
                for _ in range(patch_num):
                    samples.append((str(root / "distorted_images" / image_names[selected_idx]), labels[selected_idx]))
        super().__init__(samples, transform)


class KADID10k(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        ref_ids = sorted(list_tid_reference_ids(root / "ref_imgs", ".png.PNG"))
        image_names, targets, ref_ids_all = [], [], []
        with open(root / "dmos.csv", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                image_names.append(row["dist_img"])
                ref_ids_all.append(row["ref_img"][1:3])
                targets.append(np.asarray(float(row["dmos"])).astype(np.float32))
        labels = np.asarray(targets).astype(np.float32)
        ref_ids_all = np.asarray(ref_ids_all)
        samples = []
        for item in index:
            selected_indices = np.where(ref_ids[item] == ref_ids_all)[0].tolist()
            for selected_idx in selected_indices:
                for _ in range(patch_num):
                    samples.append((str(root / "images" / image_names[selected_idx]), labels[selected_idx]))
        super().__init__(samples, transform)


class LIVEMDDataset(IQADataset):
    def __init__(self, root, index, transform, patch_num):
        root = Path(root)
        image_paths, ref_image_names, mos_values = [], [], []
        for part, folder in zip(["Part 1", "Part 2"], ["blurjpeg", "blurnoise"]):
            part_root = root / part
            image_list = scipy.io.loadmat(part_root / "Imagelists.mat")
            dist_names = [item[0][0] for item in image_list["distimgs"]]
            score = scipy.io.loadmat(part_root / "Scores.mat")
            dmos_scores = score["DMOSscores"][0]
            for dist_name, dmos in zip(dist_names, dmos_scores):
                image_paths.append(str(part_root / folder / dist_name))
                ref_image_names.append(dist_name.split("_")[0] + ".bmp")
                mos_values.append(dmos)

        unique_refs = np.unique(ref_image_names)
        ref_image_names = np.asarray(ref_image_names)
        samples = []
        for item in index:
            selected_indices = np.where(unique_refs[item] == ref_image_names)[0].tolist()
            for selected_idx in selected_indices:
                for _ in range(patch_num):
                    samples.append((image_paths[selected_idx], mos_values[selected_idx]))
        super().__init__(samples, transform)


DATASET_CLASS_REGISTRY: Dict[str, Callable[..., data.Dataset]] = {
    "live": LIVE,
    "csiq": CSIQ,
    "tid2013": TID2013,
    "kadid-10k": KADID10k,
    "livemd": LIVEMDDataset,

    "clive": LIVEChallenge,
    "koniq-10k": KonIQ10k,
    "bid": BID,
    "spaq": SPAQ,
    "fblive": FBLIVE,

    "generated_dataset": SAQTIQADataset,
    "saqt-iqa": SAQTIQADataset,
}


def create_iqa_dataset(
    dataset_name: str,
    root: Path,
    indices: Iterable[int],
    patch_size: int,
    patch_num: int,
    is_train: bool,
    args: Any = None,
):
    dataset_name = canonical_dataset_name(dataset_name)
    root = _require_path(root, "dataset root")
    transform = build_iqa_transform(dataset_name, patch_size=patch_size, is_train=is_train)
    indices = list(indices)

    if dataset_name not in DATASET_CLASS_REGISTRY:
        supported = ", ".join(sorted(DATASET_CLASS_REGISTRY))
        raise ValueError(f"Unsupported dataset '{dataset_name}'. Supported datasets: {supported}")

    if dataset_name in {"generated_dataset", "saqt-iqa"}:
        label_path = _require_path(_get_arg(args, "label_path", None), "label_path")
        similarity_csv = _require_path(_get_arg(args, "similarity_csv", None), "similarity_csv")
        target_distortion_types = _get_arg(args, "target_distortion_types", None)
        return SAQTIQADataset(
            pretrain_path=str(root),
            index=indices,
            transform=transform,
            label_path=str(label_path),
            similarity_csv_path=str(similarity_csv),
            patch_num=patch_num,
            target_distortion_types=target_distortion_types,
        )

    if dataset_name == "clive":
        return LIVEChallenge(str(root), indices, transform, patch_num, is_train=is_train)

    dataset_class = DATASET_CLASS_REGISTRY[dataset_name]
    return dataset_class(str(root), indices, transform, patch_num)


def _get_arg(args: Any, name: str, default=None):
    return getattr(args, name, default) if args is not None else default


def _require_path(path: Optional[Path], name: str) -> Path:
    if path is None:
        raise ValueError(f"{name} is required.")
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"{name} not found: {path}")
    return path


def _first_existing_column(df: pd.DataFrame, candidates: Sequence[str]) -> str:
    for column in candidates:
        if column in df.columns:
            return column
    raise ValueError(f"None of the expected columns exist: {list(candidates)}")


def _parse_generated_distortion_name(filename: str) -> Tuple[str, str]:
    stem = Path(filename).stem
    parts = stem.rsplit("_", 2)
    if len(parts) != 3:
        raise ValueError(f"Cannot parse generated distortion name: {filename}")
    ref_stem, dist_type, _ = parts
    return f"{ref_stem}.png", dist_type


def _live_distortion_paths(root: Path, count: int) -> List[str]:
    return [str(root / f"img{idx}.bmp") for idx in range(1, count + 1)]


def _repeat_samples(root: Path, image_names, labels, indices, patch_num: int):
    samples = []
    for item in indices:
        for _ in range(patch_num):
            samples.append((str(root / image_names[item]), labels[item]))
    return samples


__all__ = [
    "BID",
    "CSIQ",
    "DATASET_CLASS_REGISTRY",
    "FBLIVE",
    "GeneratedDataset",
    "IQADataset",

    "KADID10k",
    "KonIQ10k",
    "LIVE",
    "LIVEChallenge",
    "LIVEMDDataset",
    "SAQTIQADataset",
    "SPAQ",
    "TID2013",
    "create_iqa_dataset",
    "pil_loader",
]
